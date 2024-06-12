import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Lees de Excel file
file_path = 'TagsAndNotes.xlsm'
wb = load_workbook(file_path)
sheet1 = wb['TagsAndNotes']
sheet2 = wb['Blad2']


# Converteer de sheets naar dataframes
df1 = pd.DataFrame(sheet1.values)
df2 = pd.DataFrame(sheet2.values)

# Eerste rij als kolomkoppen instellen
df1.columns = df1.iloc[0]
df2.columns = df2.iloc[0]
df1 = df1[1:]
df2 = df2[1:]

# Index instellen op de eerste kolom (unieke waarden)
df1.set_index(df1.columns[0], inplace=True)
df2.set_index(df2.columns[0], inplace=True)

# Zorg ervoor dat beide dataframes dezelfde kolommen en indexen hebben
df2 = df2.reindex(index=df1.index, columns=df1.columns)

# Vergelijk de twee dataframes
comparison = df1.compare(df2, align_axis=1)

# Voeg de verschillen toe aan sheet2 met gele markering
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for idx, row in comparison.iterrows():
    for col in comparison.columns.levels[0]:  # Now, comparison.columns is a MultiIndex
        if row[(col, 'self')] != row[(col, 'other')]:
            cell = sheet2.cell(row=df2.index.get_loc(idx)+2, column=df2.columns.get_loc(col)+1)
            cell.fill = yellow_fill
            
# Sla de gewijzigde Excel file op
output_path = 'output.xlsx'
wb.save(output_path)